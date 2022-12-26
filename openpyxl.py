from openpyxl import load_workbook

wb = load_workbook(r'C:\test\test.xlsx') #load excel
sheet = wb.active
max_row = sheet.max_row
print(max_row)

sheet['A1'] = 'School name'
sheet['B1'] = 'school url'
sheet.cell(row=max_row + 1, column=4).value = 5 #add value in cell

sheet.cell(row=1, column=2).value = "Hi" #add value in cell

wb.save(r'C:\test\test.xlsx') #save excel

#------------------CSV--------------------------------------

import csv

with open('names.csv', 'r') as csv_file:
    csv_reader = csv.reader(csv_file)

    with open('new_names.csv', 'w') as new_file:
        fieldnames = ['first_name', 'last_name'] #optional

        csv_writer = csv.writer(new_file, fieldnames=fieldnames, delimiter='\t')

        csv_writer.writeheader() #optional

        for line in csv_reader:
            csv_writer.writerow(line[3])  #row 3
